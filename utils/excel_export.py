import io
import openpyxl
from datetime import datetime, timezone, timedelta
import os

def number_to_words_vn(n):
    if n == 0:
        return "không đồng"
        
    units = ["", "nghìn", "triệu", "tỷ", "nghìn tỷ", "triệu tỷ"]
    words = ["không", "một", "hai", "ba", "bốn", "năm", "sáu", "bảy", "tám", "chín"]
    
    def read_block_3(num, full=False):
        res = ""
        hundred = num // 100
        ten = (num % 100) // 10
        unit = num % 10
        
        if full or hundred > 0:
            res += words[hundred] + " trăm "
            if ten == 0 and unit > 0:
                res += "lẻ "
                
        if ten == 1:
            res += "mười "
        elif ten > 1:
            res += words[ten] + " mươi "
            
        if unit == 1 and ten > 1:
            res += "mốt "
        elif unit == 5 and ten > 0:
            res += "lăm "
        elif unit > 0:
            res += words[unit] + " "
            
        return res.strip()

    s = ""
    block_idx = 0
    while n > 0:
        block = n % 1000
        if block > 0:
            s = read_block_3(block, n >= 1000) + " " + units[block_idx] + " " + s
        n //= 1000
        block_idx += 1
        
    s = s.strip() + " đồng"
    return s.capitalize()

def generate_payment_request_excel(debt_data):
    """
    debt_data: dict chứa thông tin nợ, bao gồm các field:
    - ten_cty, nguoi_lien_he, sdt
    - ma_hd, gia_tri_thang, don_vi_tinh
    - ky_thanh_toan, can_thu, da_thu, tien_vat
    - id
    """
    # Đường dẫn file mẫu
    template_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "Phieu_Yeu_Cau_Thanh_Toan_VHS.xlsx")
    
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # 1. Điền thông tin khách hàng và số phiếu
    ws["C9"] = debt_data.get("ten_cty", "")
    ws["C10"] = debt_data.get("nguoi_lien_he", "")
    ws["F11"] = debt_data.get("ma_hd", "")
    
    # Số phiếu
    year = debt_data.get("ky_thanh_toan", "2026")[:4]
    ws["F9"] = f"PYC-{year}-{debt_data.get('id', '')}"
    
    # Ngày tạo
    now = datetime.now(timezone.utc) + timedelta(hours=7)
    ws["F10"] = now.strftime("%d/%m/%Y")
    
    # 2. Điền thông tin dịch vụ (Dòng 14)
    # Kỳ thanh toán
    ws["D14"] = f"Tháng {debt_data.get('ky_thanh_toan', '')[-2:]}/{year}" if "-" in debt_data.get('ky_thanh_toan', '') else debt_data.get('ky_thanh_toan', '')
    
    # Thành tiền gốc (Chưa VAT)
    tien_vat = float(debt_data.get("tien_vat", 0) or 0)
    can_thu = float(debt_data.get("can_thu", 0) or 0)
    goc = can_thu - tien_vat
    
    # Update Đơn giá and Thành tiền
    ws["F14"] = goc
    ws["G14"] = goc
    
    # Clear dòng phát sinh nếu không có
    ws["F15"] = 0
    ws["G15"] = 0
    
    # Cập nhật VAT (G17)
    # File mẫu có công thức G16*0.08, ta ghi thẳng giá trị để chắc chắn
    ws["G17"] = tien_vat
    
    # Tổng cộng
    ws["G18"] = can_thu
    
    # Đọc số thành chữ
    words = number_to_words_vn(int(can_thu))
    ws["C19"] = words
    
    # Nội dung CK
    ten_cty = debt_data.get("ten_cty", "")
    short_name = ten_cty[:20] if ten_cty else "Khach hang"
    ws["C25"] = f"VHS {short_name} thanh toan"
    
    # Lưu vào buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue(), f"PYC_{debt_data.get('ma_hd', '')}_{debt_data.get('ky_thanh_toan', '')}.xlsx"
